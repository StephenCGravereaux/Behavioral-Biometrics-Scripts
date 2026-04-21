#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Behavioral Biometric Analysis Pipeline
Paper: "Behavioral Biometric Analysis for Task and Keystroke inference for task recognition"

Three-stage pipeline:
  Stage 1 â€” Attack Demonstration: baseline classifiers on clean data
  Stage 2 â€” IBM DP Defense: Laplace mechanism via diffprivlib
  Stage 3 â€” Privacy-Utility Evaluation: epsilon sweep {0.05, 0.1, 0.5, 1.0, 2.0}

WISDM task   : binary typing-task inference (F+Q vs. all others), 182 smartwatch features
Keystroke100 : closed-set 100-user identification, 13 tabular features
"""

import os, io, sys, warnings
from datetime import datetime
import numpy as np
import pandas as pd
from scipy.io import arff
from sklearn.preprocessing import StandardScaler
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import (accuracy_score, precision_score, recall_score,
                              f1_score, roc_auc_score, roc_curve)
from sklearn.model_selection import StratifiedShuffleSplit, StratifiedKFold
from sklearn.base import clone
try:
    from diffprivlib.mechanisms import Laplace as DPLaplace
except ImportError as exc:
    raise ImportError(
        "Missing dependency 'diffprivlib'. Install with: pip install diffprivlib"
    ) from exc

warnings.filterwarnings('ignore')

SEED = 42
np.random.seed(SEED)

# â”€â”€â”€ Dataset paths â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_ROOT = os.path.dirname(os.path.abspath(__file__))


def _resolve_dataset_base():
    """
    Resolve dataset root.
    Priority:
      1) SMARTWATCH_DATASETS_DIR env var
      2) ./Datasets next to this script
    """
    candidates = []
    env_base = os.environ.get('SMARTWATCH_DATASETS_DIR')
    if env_base:
        candidates.append(env_base)
    candidates.append(os.path.join(_ROOT, 'Datasets'))

    for base in candidates:
        if base and os.path.isdir(base):
            return base

    checked = '\n'.join(f'  - {c}' for c in candidates if c)
    raise FileNotFoundError(
        "Could not find dataset directory.\n"
        f"Checked:\n{checked}\n"
        "Set SMARTWATCH_DATASETS_DIR or place 'Datasets' next to pipeline.py."
    )


_BASE = _resolve_dataset_base()
WISDM_WATCH = os.path.join(
    _BASE,
    'wisdm+smartphone+and+smartwatch+activity+and+biometrics+dataset',
    'wisdm-dataset', 'wisdm-dataset', 'arff_files', 'watch')
KEYSTROKE_DIR = os.path.join(_BASE, 'keystrokes', 'keystroke100')

EPSILONS = [0.05, 0.1, 0.5, 1.0, 2.0]

# DP strategy:
#   "fixed"    -> original per-feature fixed epsilon sweep (EPSILONS)
#   "adaptive" -> typing-risk-triggered epsilon schedule for WISDM
DP_STRATEGY = os.environ.get('SMARTWATCH_DP_STRATEGY', 'adaptive').strip().lower()
if DP_STRATEGY not in {'fixed', 'adaptive'}:
    DP_STRATEGY = 'adaptive'

# Adaptive DP config (WISDM)
ADP_TAU = float(os.environ.get('SMARTWATCH_ADP_TAU', 0.60))
ADP_EPS_NONTYPING = float(os.environ.get('SMARTWATCH_ADP_EPS_NONTYPING', 2.0))
ADP_EPS_TYPING = [0.05, 0.1, 0.2, 0.5, 1.0]
# ADP policy:
#   "binary" -> hard switch at tau (legacy behavior)
#   "smooth" -> continuous epsilon mapping from risk + bounded random jitter
ADP_POLICY = os.environ.get('SMARTWATCH_ADP_POLICY', 'smooth').strip().lower()
if ADP_POLICY not in {'binary', 'smooth'}:
    ADP_POLICY = 'smooth'
ADP_JITTER_FRAC = float(os.environ.get('SMARTWATCH_ADP_JITTER_FRAC', 0.15))
ADP_JITTER_FRAC = float(np.clip(ADP_JITTER_FRAC, 0.0, 0.49))


def validate_input_paths():
    """Fail fast with clear messages if input folders are missing."""
    required_dirs = {
        'WISDM watch accel ARFF dir': os.path.join(WISDM_WATCH, 'accel'),
        'WISDM watch gyro ARFF dir': os.path.join(WISDM_WATCH, 'gyro'),
        'Keystroke100 dir': KEYSTROKE_DIR,
    }
    missing = [f'{name}: {path}' for name, path in required_dirs.items()
               if not os.path.isdir(path)]
    if missing:
        raise FileNotFoundError(
            'Missing required input directories:\n  - ' + '\n  - '.join(missing)
        )

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# DATA LOADING
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _load_arff_dir(dirpath):
    """Load all .arff files from dirpath, return concatenated DataFrame."""
    frames = []
    for fname in sorted(os.listdir(dirpath)):
        if not fname.endswith('.arff'):
            continue
        with open(os.path.join(dirpath, fname), 'rb') as fh:
            text = fh.read().decode('utf-8', errors='replace')
        data, _ = arff.loadarff(io.StringIO(text))
        df = pd.DataFrame(data)
        # Decode bytes columns
        for col in df.select_dtypes(['object']).columns:
            df[col] = df[col].apply(
                lambda v: v.decode('utf-8', errors='replace').strip()
                if isinstance(v, bytes) else str(v).strip())
        frames.append(df)
    if not frames:
        raise FileNotFoundError(f'No .arff files found in: {dirpath}')
    return pd.concat(frames, ignore_index=True)


def load_wisdm():
    """
    Load watch accel + gyro ARFF files.
    Merge by (subject, activity, window_index) -> 182 numeric features.
    Binary label: F (typing) or Q (writing) = 1, all others = 0.

    Returns
    -------
    X        : (N, 182) float64
    y        : (N,) int   binary typing label
    subjects : (N,) str   subject IDs (for disjoint splitting)
    activities: (N,) str  activity codes (for utility model)
    """
    print("Loading WISDM watch accelerometer â€¦")
    acc = _load_arff_dir(os.path.join(WISDM_WATCH, 'accel'))
    print(f"  accel shape : {acc.shape}")

    print("Loading WISDM watch gyroscope â€¦")
    gyr = _load_arff_dir(os.path.join(WISDM_WATCH, 'gyro'))
    print(f"  gyro  shape : {gyr.shape}")

    # Column names in the ARFF are quoted, e.g. "ACTIVITY", "class"
    ACT_COL   = '"ACTIVITY"'
    CLS_COL   = '"class"'

    for df in (acc, gyr):
        df['_subj'] = df[CLS_COL].str.strip('"')
        df['_act']  = df[ACT_COL].str.strip('"')

    acc.sort_values(['_subj', '_act'], inplace=True)
    gyr.sort_values(['_subj', '_act'], inplace=True)
    acc.reset_index(drop=True, inplace=True)
    gyr.reset_index(drop=True, inplace=True)

    acc['_win'] = acc.groupby(['_subj', '_act']).cumcount()
    gyr['_win'] = gyr.groupby(['_subj', '_act']).cumcount()

    _meta = {ACT_COL, CLS_COL, '_subj', '_act', '_win'}
    acc_feat = [c for c in acc.columns if c not in _meta]
    gyr_feat = [c for c in gyr.columns if c not in _meta]

    a = acc[['_subj', '_act', '_win'] + acc_feat].rename(
        columns={c: f'A_{c}' for c in acc_feat})
    g = gyr[['_subj', '_act', '_win'] + gyr_feat].rename(
        columns={c: f'G_{c}' for c in gyr_feat})

    merged = pd.merge(a, g, on=['_subj', '_act', '_win'], how='inner')
    print(f"  merged shape: {merged.shape}")

    feat_cols = [c for c in merged.columns if c.startswith(('A_', 'G_'))]
    X          = merged[feat_cols].values.astype(np.float64)
    activities = merged['_act'].values
    subjects   = merged['_subj'].values
    y          = np.where(np.isin(activities, ['F', 'Q']), 1, 0)

    print(f"  features={X.shape[1]}  pos(F+Q)={y.sum()}  neg={( y==0).sum()}")
    return X, y, subjects, activities


def load_keystroke100():
    """
    Load Keystroke100: 13 features per sample.
      7  latency features  (from latency.txt rows)
      6  pressure-summary  (mean, std, max, min, peak-to-peak, duration)

    Returns
    -------
    X : (1000, 13) float64
    y : (1000,)   int  user label 0-99
    """
    X_all, y_all = [], []
    user_dirs = sorted(
        [d for d in os.listdir(KEYSTROKE_DIR)
         if os.path.isdir(os.path.join(KEYSTROKE_DIR, d))],
        key=lambda d: int(d.replace('user', '')))
    if not user_dirs:
        raise FileNotFoundError(f'No user folders found in: {KEYSTROKE_DIR}')

    for uid, udir in enumerate(user_dirs):
        upath = os.path.join(KEYSTROKE_DIR, udir)
        lat = np.loadtxt(os.path.join(upath, 'latency.txt'))   # (10, 7)
        # Use pandas to handle ragged rows gracefully
        prs_df = pd.read_csv(os.path.join(upath, 'pressure.txt'),
                             sep=r'\s+', header=None, engine='python')
        prs = prs_df.fillna(prs_df.median()).values                 # (T, â‰¤10)

        if lat.ndim == 1:
            lat = lat.reshape(1, -1)
        if prs.ndim == 1:
            prs = prs.reshape(-1, 1)

        # columns = samples
        prs_T = prs.T                   # (10, T)
        n = min(lat.shape[0], prs_T.shape[0])
        p = prs_T[:n]
        T = p.shape[1]

        prs_feats = np.column_stack([
            p.mean(axis=1),
            p.std(axis=1),
            p.max(axis=1),
            p.min(axis=1),
            p.max(axis=1) - p.min(axis=1),   # peak-to-peak
            np.full(n, float(T)),             # duration (time steps)
        ])

        feats = np.hstack([lat[:n], prs_feats])   # (n, 13)
        X_all.append(feats)
        y_all.extend([uid] * n)

    X = np.vstack(X_all)
    y = np.array(y_all, dtype=int)
    print(f"Keystroke100: shape={X.shape}  classes={len(np.unique(y))}")
    return X, y


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SPLITTING
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def subject_disjoint_split(X, y, subjects, activities=None,
                           train_f=0.70, val_f=0.15, seed=SEED):
    """70/15/15 subject-disjoint split."""
    rng  = np.random.RandomState(seed)
    uniq = rng.permutation(np.unique(subjects))
    n    = len(uniq)
    n_tr = int(n * train_f)
    n_vl = int(n * val_f)

    tr  = set(uniq[:n_tr])
    vl  = set(uniq[n_tr:n_tr + n_vl])
    te  = set(uniq[n_tr + n_vl:])

    def idx(s): return np.where([x in s for x in subjects])[0]
    i_tr, i_vl, i_te = idx(tr), idx(vl), idx(te)

    if activities is not None:
        return (X[i_tr], y[i_tr], activities[i_tr],
                X[i_vl], y[i_vl], activities[i_vl],
                X[i_te], y[i_te], activities[i_te])
    return (X[i_tr], y[i_tr],
            X[i_vl], y[i_vl],
            X[i_te], y[i_te])


def stratified_split(X, y, train_f=0.70, val_f=0.15, seed=SEED):
    """Stratified 70/15/15 split."""
    sss1 = StratifiedShuffleSplit(1, test_size=round(1 - train_f, 4), random_state=seed)
    tr, rest = next(sss1.split(X, y))
    rel = val_f / (1 - train_f)
    sss2 = StratifiedShuffleSplit(1, test_size=round(1 - rel, 4), random_state=seed)
    vl_r, te_r = next(sss2.split(X[rest], y[rest]))
    return (X[tr], y[tr],
            X[rest[vl_r]], y[rest[vl_r]],
            X[rest[te_r]], y[rest[te_r]])


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# DIFFERENTIAL PRIVACY (Laplace mechanism, IBM diffprivlib approach)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def laplace_dp(X_ref, X_target, epsilon):
    """
    Per-feature Laplace DP perturbation via IBM diffprivlib.
    Bounds [Lj, Uj] computed from X_ref (training split only).
    DPLaplace mechanism: scale bj = sensitivity_j / epsilon.
    Applied vectorised (equivalent to calling mech.randomise per sample).
    """
    lo   = X_ref.min(axis=0)
    hi   = X_ref.max(axis=0)
    sens = hi - lo
    Xc   = np.clip(X_target, lo, hi)
    out  = Xc.copy()
    for j in range(Xc.shape[1]):
        s = float(sens[j])
        if s <= 0:
            continue
        mech = DPLaplace(epsilon=epsilon, sensitivity=s)
        # mech.randomise(v) = v + Laplace(0, s/epsilon); apply vectorised
        out[:, j] += np.random.laplace(0.0, s / mech.epsilon, Xc.shape[0])
    return out


def laplace_dp_adaptive(X_ref, X_target, eps_row):
    """
    Row-adaptive Laplace DP perturbation.
    Each row i uses epsilon eps_row[i] across all features.
    """
    eps = np.asarray(eps_row, dtype=np.float64).reshape(-1)
    if eps.shape[0] != X_target.shape[0]:
        raise ValueError(
            f'eps_row length {eps.shape[0]} does not match rows {X_target.shape[0]}')

    # Avoid invalid or near-zero epsilon values.
    eps = np.clip(eps, 1e-6, None)

    lo   = X_ref.min(axis=0)
    hi   = X_ref.max(axis=0)
    sens = hi - lo
    Xc   = np.clip(X_target, lo, hi)
    out  = Xc.copy()

    # Vectorised noise draw: scale_ij = sensitivity_j / eps_i
    scales = sens[np.newaxis, :] / eps[:, np.newaxis]
    noise = np.random.laplace(loc=0.0, scale=scales, size=Xc.shape)
    out += noise
    return out


def _typing_risk_scores(clf, scaler, scaled, X):
    """Return typing-risk scores in [0,1] from the stage-1 typing model."""
    Xin = scaler.transform(X) if scaled else X
    if hasattr(clf, 'predict_proba'):
        return clf.predict_proba(Xin)[:, 1]
    if hasattr(clf, 'decision_function'):
        z = clf.decision_function(Xin)
        return 1.0 / (1.0 + np.exp(-z))
    return clf.predict(Xin).astype(float)


def _adaptive_eps_schedule(risk_scores, eps_typing, eps_nontyping, tau,
                           policy='binary', jitter_frac=0.0, rng=None):
    """
    Adaptive epsilon schedule.
      - binary: hard switch at tau
      - smooth: continuous map eps(risk) between [eps_typing, eps_nontyping]
                with bounded multiplicative jitter to reduce regime artifacts
    Returns (eps_row, mask_typing) where mask_typing is risk>=tau.
    """
    risk = np.asarray(risk_scores, dtype=np.float64).reshape(-1)
    risk = np.clip(risk, 0.0, 1.0)
    eps_t = float(eps_typing)
    eps_n = float(eps_nontyping)
    mask_typing = risk >= tau

    if policy == 'smooth':
        # High risk -> lower epsilon (more noise), low risk -> higher epsilon.
        eps_row = eps_t + (eps_n - eps_t) * (1.0 - risk)
        if jitter_frac > 0:
            if rng is None:
                rng = np.random
            jitter = rng.uniform(-jitter_frac, jitter_frac, size=risk.shape[0])
            eps_row = eps_row * (1.0 + jitter)
            lo = min(eps_t, eps_n)
            hi = max(eps_t, eps_n)
            eps_row = np.clip(eps_row, lo, hi)
    else:
        eps_row = np.where(mask_typing, eps_t, eps_n)

    return eps_row, mask_typing


def _oof_typing_risk_scores(clf_template, X, y, scaled, n_splits=5, seed=SEED):
    """
    Out-of-fold typing-risk scores for X.
    Prevents in-sample gate leakage when adaptive policy is trained/evaluated.
    """
    y = np.asarray(y)
    oof = np.zeros(len(y), dtype=np.float64)
    skf = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=seed)

    for tr, va in skf.split(X, y):
        clf = clone(clf_template)
        if scaled:
            sc = StandardScaler().fit(X[tr])
            xtr = sc.transform(X[tr])
            xva = sc.transform(X[va])
        else:
            sc = None
            xtr = X[tr]
            xva = X[va]
        clf.fit(xtr, y[tr])

        if hasattr(clf, 'predict_proba'):
            oof[va] = clf.predict_proba(xva)[:, 1]
        elif hasattr(clf, 'decision_function'):
            z = clf.decision_function(xva)
            oof[va] = 1.0 / (1.0 + np.exp(-z))
        else:
            oof[va] = clf.predict(xva).astype(float)

    return oof


def _safe_corr(a, b):
    """Pearson correlation with guard against degenerate vectors."""
    a = np.asarray(a, dtype=np.float64).reshape(-1)
    b = np.asarray(b, dtype=np.float64).reshape(-1)
    if a.size != b.size or a.size == 0:
        return float('nan')
    if np.all(a == a[0]) or np.all(b == b[0]):
        return float('nan')
    return float(np.corrcoef(a, b)[0, 1])


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# METRIC HELPERS
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def binary_metrics(clf, X, y):
    yp = clf.predict(X)
    return binary_metrics_from_pred(y, yp, clf=clf, X=X)


def binary_metrics_from_pred(y, yp, clf=None, X=None):
    """Binary metrics from precomputed predictions; optionally include ROC-AUC."""
    m  = dict(
        acc       = accuracy_score(y, yp),
        precision = precision_score(y, yp, zero_division=0),
        recall    = recall_score(y, yp, zero_division=0),
        f1        = f1_score(y, yp, zero_division=0),
    )
    if clf is not None and X is not None and hasattr(clf, 'predict_proba'):
        m['roc_auc'] = roc_auc_score(y, clf.predict_proba(X)[:, 1])
    elif clf is not None and X is not None and hasattr(clf, 'decision_function'):
        m['roc_auc'] = roc_auc_score(y, clf.decision_function(X))
    else:
        m['roc_auc'] = float('nan')
    return m


def top_k_acc(y, proba, k):
    top = np.argsort(proba, axis=1)[:, -k:]
    return float(np.mean([y[i] in top[i] for i in range(len(y))]))


def _eer(y_bin, scores):
    fpr, tpr, _ = roc_curve(y_bin, scores)
    fnr = 1 - tpr
    i   = np.argmin(np.abs(fpr - fnr))
    return float((fpr[i] + fnr[i]) / 2)


def multiclass_metrics(clf, X, y, classes):
    yp    = clf.predict(X)
    proba = clf.predict_proba(X) if hasattr(clf, 'predict_proba') else None
    m = dict(
        top1      = accuracy_score(y, yp),
        top5      = top_k_acc(y, proba, 5) if proba is not None else float('nan'),
        precision = precision_score(y, yp, average='weighted', zero_division=0),
        recall    = recall_score(y, yp, average='weighted', zero_division=0),
        f1        = f1_score(y, yp, average='weighted', zero_division=0),
    )
    if proba is not None:
        eers = [_eer((y == c).astype(int), proba[:, i])
                for i, c in enumerate(classes)
                if 0 < (y == c).sum() < len(y)]
        m['mean_eer'] = float(np.mean(eers)) if eers else float('nan')
    else:
        m['mean_eer'] = float('nan')
    return m


def _fmt_binary(name, split, m):
    print(f"  {name:<22} {split:<5}  "
          f"acc={m['acc']:.4f}  prec={m['precision']:.4f}  "
          f"rec={m['recall']:.4f}  f1={m['f1']:.4f}  auc={m['roc_auc']:.4f}")


def _fmt_mc(name, split, m):
    print(f"  {name:<15} {split:<5}  "
          f"top1={m['top1']:.4f}  top5={m['top5']:.4f}  "
          f"prec={m['precision']:.4f}  rec={m['recall']:.4f}  "
          f"f1={m['f1']:.4f}  eer={m['mean_eer']:.4f}")


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# STAGE 1 â€” WISDM: Typing Task Inference
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def stage1_wisdm():
    print("\n" + "=" * 72)
    print("STAGE 1 â€” WISDM: Typing Task Inference  (F+Q = 1, rest = 0)")
    print("=" * 72)

    X, y, subjects, activities = load_wisdm()
    (X_tr, y_tr, act_tr,
     X_vl, y_vl, act_vl,
     X_te, y_te, act_te) = subject_disjoint_split(X, y, subjects, activities)

    print(f"\nSplit (subject-disjoint 70/15/15):")
    print(f"  train : {len(y_tr):5d}  pos={y_tr.sum():4d}  neg={(y_tr==0).sum():4d}")
    print(f"  val   : {len(y_vl):5d}  pos={y_vl.sum():4d}  neg={(y_vl==0).sum():4d}")
    print(f"  test  : {len(y_te):5d}  pos={y_te.sum():4d}  neg={(y_te==0).sum():4d}")

    sc     = StandardScaler().fit(X_tr)
    Xs_tr  = sc.transform(X_tr)
    Xs_vl  = sc.transform(X_vl)
    Xs_te  = sc.transform(X_te)

    models = {
        'LogisticRegression': (LogisticRegression(max_iter=1000, random_state=SEED), True),
        'RandomForest'      : (RandomForestClassifier(n_estimators=200, n_jobs=1,
                                                       random_state=SEED), False),
        'SVM_RBF'           : (SVC(kernel='rbf', probability=True, random_state=SEED), True),
    }

    val_scores = {}
    print("\nBaseline Results (no DP):")
    for name, (clf, scaled) in models.items():
        clf.fit(Xs_tr if scaled else X_tr, y_tr)
        m_vl = binary_metrics(clf, Xs_vl if scaled else X_vl, y_vl)
        m_te = binary_metrics(clf, Xs_te if scaled else X_te, y_te)
        val_scores[name] = (m_vl, m_te, scaled)
        _fmt_binary(name, 'val',  m_vl)
        _fmt_binary(name, 'test', m_te)

    # Best on val F1; tiebreaker ROC-AUC
    best = max(val_scores, key=lambda n: (val_scores[n][0]['f1'],
                                          val_scores[n][0]['roc_auc']))
    print(f"\nBest model (val F1): {best}")

    # Refit best on train + val
    X_tv   = np.vstack([X_tr, X_vl])
    y_tv   = np.hstack([y_tr, y_vl])
    act_tv = np.hstack([act_tr, act_vl])
    _, _, scaled = val_scores[best]

    sc2      = StandardScaler().fit(X_tv)
    X_tv_in  = sc2.transform(X_tv) if scaled else X_tv
    X_te_in  = sc2.transform(X_te) if scaled else X_te

    BestCls  = models[best][0].__class__
    best_clf = BestCls(**models[best][0].get_params())
    best_clf.fit(X_tv_in, y_tv)

    m_final = binary_metrics(best_clf, X_te_in, y_te)
    print(f"\nFinal test ({best}, refitted on train+val):")
    _fmt_binary(best, 'test', m_final)

    return dict(
        best_clf=best_clf, best_name=best, scaled=scaled, scaler=sc2,
        X_tv=X_tv, y_tv=y_tv, act_tv=act_tv,
        X_te=X_te, y_te=y_te, act_te=act_te,
        baseline=m_final,
        model_metrics=val_scores,   # {name: (m_val, m_test, scaled)}
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# STAGE 1 â€” Keystroke100: User Identification
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def stage1_keystroke():
    print("\n" + "=" * 72)
    print("STAGE 1 â€” Keystroke100: Closed-Set User Identification (100 classes)")
    print("=" * 72)

    X, y = load_keystroke100()
    X_tr, y_tr, X_vl, y_vl, X_te, y_te = stratified_split(X, y)
    classes = np.unique(y)

    print(f"\nSplit (stratified 70/15/15):")
    print(f"  train={len(y_tr)}  val={len(y_vl)}  test={len(y_te)}  classes={len(classes)}")

    sc    = StandardScaler().fit(X_tr)
    Xs_tr = sc.transform(X_tr)
    Xs_vl = sc.transform(X_vl)
    Xs_te = sc.transform(X_te)

    models = {
        'kNN'         : (KNeighborsClassifier(n_neighbors=5, n_jobs=1), True),
        'RandomForest': (RandomForestClassifier(n_estimators=200, n_jobs=1,
                                                random_state=SEED), False),
        'SVM_RBF'     : (SVC(kernel='rbf', probability=True, random_state=SEED), True),
    }

    val_scores = {}
    print("\nBaseline Results (no DP):")
    for name, (clf, scaled) in models.items():
        clf.fit(Xs_tr if scaled else X_tr, y_tr)
        m_vl = multiclass_metrics(clf, Xs_vl if scaled else X_vl, y_vl, classes)
        m_te = multiclass_metrics(clf, Xs_te if scaled else X_te, y_te, classes)
        val_scores[name] = (m_vl, m_te, scaled)
        _fmt_mc(name, 'val',  m_vl)
        _fmt_mc(name, 'test', m_te)

    # Best on val top-1; tiebreaker top-5
    best = max(val_scores, key=lambda n: (val_scores[n][0]['top1'],
                                          val_scores[n][0]['top5']))
    print(f"\nBest model (val Top-1): {best}")

    X_tv  = np.vstack([X_tr, X_vl])
    y_tv  = np.hstack([y_tr, y_vl])
    _, _, scaled = val_scores[best]

    sc2      = StandardScaler().fit(X_tv)
    X_tv_in  = sc2.transform(X_tv) if scaled else X_tv
    X_te_in  = sc2.transform(X_te) if scaled else X_te

    BestCls  = models[best][0].__class__
    best_clf = BestCls(**models[best][0].get_params())
    best_clf.fit(X_tv_in, y_tv)

    m_final = multiclass_metrics(best_clf, X_te_in, y_te, classes)
    print(f"\nFinal test ({best}, refitted on train+val):")
    _fmt_mc(best, 'test', m_final)

    return dict(
        best_clf=best_clf, best_name=best, scaled=scaled, scaler=sc2,
        X_tv=X_tv, y_tv=y_tv,
        X_te=X_te, y_te=y_te,
        classes=classes,
        baseline=m_final,
        model_metrics=val_scores,   # {name: (m_val, m_test, scaled)}
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# STAGE 2+3 â€” WISDM: DP Defense + Privacy-Utility Tradeoff
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _balanced_clf(clf):
    """Return a copy of clf with class_weight='balanced' if supported."""
    params = clf.get_params()
    if 'class_weight' in params:
        params['class_weight'] = 'balanced'
        return clf.__class__(**params)
    return clf.__class__(**params)   # kNN â€” no class_weight, returns unmodified copy


def _utility_accuracy(X_train, act_train, X_test, act_test):
    """Train a utility model on X_train and score it on X_test."""
    util_clf = RandomForestClassifier(n_estimators=200, n_jobs=-1, random_state=SEED)
    util_clf.fit(X_train, act_train)
    return accuracy_score(act_test, util_clf.predict(X_test))


def dp_eval_wisdm(res, mode=None):
    eval_mode = (mode or DP_STRATEGY).strip().lower()
    if eval_mode not in {'fixed', 'adaptive'}:
        eval_mode = DP_STRATEGY

    print("\n" + "=" * 72)
    if eval_mode == 'adaptive':
        print("STAGE 2+3 â€” WISDM: Adaptive Laplace DP (typing-triggered) + Privacy-Utility Sweep")
    else:
        print("STAGE 2+3 â€” WISDM: Laplace DP  +  Privacy-Utility Sweep")
    print("=" * 72)

    X_tv, y_tv, act_tv = res['X_tv'], res['y_tv'], res['act_tv']
    X_te, y_te, act_te = res['X_te'], res['y_te'], res['act_te']
    clf, sc, scaled = res['best_clf'], res['scaler'], res['scaled']

    nt_tv = np.where(y_tv == 0)[0]
    nt_te = np.where(y_te == 0)[0]

    clean_atk  = res['baseline']['acc']
    clean_f1   = res['baseline']['f1']
    clean_rec  = res['baseline']['recall']
    clean_util = _utility_accuracy(X_tv[nt_tv], act_tv[nt_tv], X_te[nt_te], act_te[nt_te])
    print(f"Clean attack: acc={clean_atk:.4f}  typing-recall={clean_rec:.4f}  "
          f"typing-F1={clean_f1:.4f}")
    print(f"Clean utility accuracy: {clean_util:.4f}  (non-typing activity recognition)")

    rows = []
    if eval_mode == 'adaptive':
        print(f"Adaptive config: policy={ADP_POLICY}, tau={ADP_TAU:.2f}, "
              f"eps_non_typing={ADP_EPS_NONTYPING:.2f}, eps_typing_sweep={ADP_EPS_TYPING}, "
              f"jitter_frac={ADP_JITTER_FRAC:.2f}")
        # Gate policy scores:
        # - TV uses out-of-fold scores to avoid in-sample leakage artifacts.
        # - TE uses the deployed full model score path.
        risk_tv_oof = _oof_typing_risk_scores(clf, X_tv, y_tv, scaled, n_splits=5, seed=SEED)
        risk_tv_in  = _typing_risk_scores(clf, sc, scaled, X_tv)
        risk_te = _typing_risk_scores(clf, sc, scaled, X_te)
        gate_pos_rate = float((risk_te >= ADP_TAU).mean())
        gate_tv_in = (risk_tv_in >= ADP_TAU).astype(int)
        gate_tv_oof = (risk_tv_oof >= ADP_TAU).astype(int)
        gate_te_ref = (risk_te >= ADP_TAU).astype(int)
        print(f"Typing gate positive-rate on test: {gate_pos_rate:.4f}")
        print(f"Gate quality (TV in-sample):  prec={precision_score(y_tv, gate_tv_in, zero_division=0):.4f}  "
              f"rec={recall_score(y_tv, gate_tv_in, zero_division=0):.4f}  "
              f"f1={f1_score(y_tv, gate_tv_in, zero_division=0):.4f}")
        print(f"Gate quality (TV out-of-fold): prec={precision_score(y_tv, gate_tv_oof, zero_division=0):.4f}  "
              f"rec={recall_score(y_tv, gate_tv_oof, zero_division=0):.4f}  "
              f"f1={f1_score(y_tv, gate_tv_oof, zero_division=0):.4f}")
        print(f"Gate quality (TE):             prec={precision_score(y_te, gate_te_ref, zero_division=0):.4f}  "
              f"rec={recall_score(y_te, gate_te_ref, zero_division=0):.4f}  "
              f"f1={f1_score(y_te, gate_te_ref, zero_division=0):.4f}")

        hdr = (f"\n{'eps_t':>6}  {'eps_n':>6}  {'epsμ':>8}  {'epsσ':>8}  "
               f"{'ct-F1':>8}  {'aw-F1':>8}  {'gate-F1':>8}  "
               f"{'aw~gate':>8}  {'PrivGain':>10}  {'UtilCT':>8}  {'UtilAW':>8}")
        print(hdr)
        print("-" * 122)

        for eps_t in ADP_EPS_TYPING:
            seed_base = SEED + int(round(float(eps_t) * 10000))
            rng_tv = np.random.RandomState(seed_base + 1)
            rng_te = np.random.RandomState(seed_base + 2)
            eps_tv, gate_tv = _adaptive_eps_schedule(
                risk_tv_oof, eps_t, ADP_EPS_NONTYPING, ADP_TAU,
                policy=ADP_POLICY, jitter_frac=ADP_JITTER_FRAC, rng=rng_tv)
            eps_te, gate_te = _adaptive_eps_schedule(
                risk_te, eps_t, ADP_EPS_NONTYPING, ADP_TAU,
                policy=ADP_POLICY, jitter_frac=ADP_JITTER_FRAC, rng=rng_te)
            gate_te = gate_te.astype(int)

            # Perturb test set (bounds from training split)
            X_te_dp = laplace_dp_adaptive(X_tv, X_te, eps_te)
            te_in   = sc.transform(X_te_dp) if scaled else X_te_dp

            # Perturb train+val (for attacker-aware variants)
            X_tv_dp = laplace_dp_adaptive(X_tv, X_tv, eps_tv)
            sc_aw   = StandardScaler().fit(X_tv_dp)
            tv_aw   = sc_aw.transform(X_tv_dp) if scaled else X_tv_dp
            te_aw   = sc_aw.transform(X_te_dp) if scaled else X_te_dp

            # Attack 1: clean-trained on noisy test
            m_ct = binary_metrics(clf, te_in, y_te)

            # Attack 2: attacker-aware — standard
            clf_aw = clf.__class__(**clf.get_params())
            clf_aw.fit(tv_aw, y_tv)
            yp_aw = clf_aw.predict(te_aw)
            m_aw = binary_metrics_from_pred(y_te, yp_aw, clf=clf_aw, X=te_aw)

            # Attack 3: attacker-aware + class_weight='balanced'
            clf_bal = _balanced_clf(clf)
            clf_bal.fit(tv_aw, y_tv)
            yp_bal = clf_bal.predict(te_aw)
            m_bal = binary_metrics_from_pred(y_te, yp_bal, clf=clf_bal, X=te_aw)

            # Artifact diagnostics: gate-only baseline + attacker alignment to gate
            m_gate = binary_metrics_from_pred(y_te, gate_te)
            aw_gate_agree = float(np.mean(yp_aw == gate_te))
            aw_gate_corr = _safe_corr(yp_aw, gate_te)

            # Utility: compare clean-trained vs utility-aware utility models
            nutil = _utility_accuracy(X_tv[nt_tv], act_tv[nt_tv], X_te_dp[nt_te], act_te[nt_te])
            nutil_aw = _utility_accuracy(X_tv_dp[nt_tv], act_tv[nt_tv], X_te_dp[nt_te], act_te[nt_te])
            priv_g = clean_atk - m_ct['acc']
            u_loss = clean_util - nutil
            u_loss_aw = clean_util - nutil_aw

            # Gate diagnostics (proxy for context detector quality)
            pos = y_te == 1
            neg = y_te == 0
            gate_rec = m_gate['recall']
            gate_fpr = float(gate_te[neg].mean()) if neg.any() else float('nan')
            eff_eps = float(eps_te.mean())
            eff_eps_std = float(eps_te.std())

            rows.append(dict(
                epsilon=eps_t,  # x-axis value in existing plotting/export code
                ct_acc=m_ct['acc'],    ct_recall=m_ct['recall'],  ct_f1=m_ct['f1'],
                aw_acc=m_aw['acc'],    aw_recall=m_aw['recall'],  aw_f1=m_aw['f1'],
                bal_acc=m_bal['acc'],  bal_recall=m_bal['recall'], bal_f1=m_bal['f1'],
                priv_gain=priv_g, util_noisy=nutil, util_loss=u_loss,
                util_noisy_aw=nutil_aw, util_loss_aw=u_loss_aw,
                policy_mode='adaptive',
                eps_typing=float(eps_t),
                eps_non_typing=float(ADP_EPS_NONTYPING),
                tau=float(ADP_TAU),
                eff_eps=eff_eps,
                eff_eps_std=eff_eps_std,
                adp_policy=ADP_POLICY,
                adp_jitter_frac=float(ADP_JITTER_FRAC),
                gate_acc=m_gate['acc'],
                gate_precision=m_gate['precision'],
                gate_recall=gate_rec,
                gate_f1=m_gate['f1'],
                gate_fpr=gate_fpr,
                aw_gate_agree=aw_gate_agree,
                aw_gate_corr=aw_gate_corr,
            ))
            print(f"{eps_t:>6.2f}  {ADP_EPS_NONTYPING:>6.2f}  "
                  f"{eff_eps:>8.4f}  {eff_eps_std:>8.4f}  "
                  f"{m_ct['f1']:>8.4f}  {m_aw['f1']:>8.4f}  {m_gate['f1']:>8.4f}  "
                  f"{aw_gate_agree:>8.4f}  {priv_g:>10.4f}  {nutil:>8.4f}  {nutil_aw:>8.4f}")
    else:
        hdr = (f"\n{'eps':>6}  {'ct-Acc':>8}  {'ct-Rec':>8}  {'ct-F1':>8}  "
               f"{'aw-Acc':>8}  {'aw-Rec':>8}  {'aw-F1':>8}  "
               f"{'bal-Acc':>9}  {'bal-F1':>8}  "
               f"{'PrivGain':>10}  {'UtilCT':>8}  {'UtilAW':>8}  {'UtilLoss':>10}")
        print(hdr)
        print("-" * 122)

        for eps in EPSILONS:
            # Perturb test set (bounds from training split)
            X_te_dp = laplace_dp(X_tv, X_te, eps)
            te_in   = sc.transform(X_te_dp) if scaled else X_te_dp

            # Perturb train+val (for attacker-aware variants)
            X_tv_dp = laplace_dp(X_tv, X_tv, eps)
            sc_aw   = StandardScaler().fit(X_tv_dp)
            tv_aw   = sc_aw.transform(X_tv_dp) if scaled else X_tv_dp
            te_aw   = sc_aw.transform(X_te_dp) if scaled else X_te_dp

            # Attack 1: clean-trained on noisy test
            m_ct = binary_metrics(clf, te_in, y_te)

            # Attack 2: attacker-aware — standard (collapses to majority class)
            clf_aw = clf.__class__(**clf.get_params())
            clf_aw.fit(tv_aw, y_tv)
            m_aw = binary_metrics(clf_aw, te_aw, y_te)

            # Attack 3: attacker-aware + class_weight='balanced' (collapse prevention)
            clf_bal = _balanced_clf(clf)
            clf_bal.fit(tv_aw, y_tv)
            m_bal = binary_metrics(clf_bal, te_aw, y_te)

            # Utility: compare clean-trained vs utility-aware utility models
            nutil = _utility_accuracy(X_tv[nt_tv], act_tv[nt_tv], X_te_dp[nt_te], act_te[nt_te])
            nutil_aw = _utility_accuracy(X_tv_dp[nt_tv], act_tv[nt_tv], X_te_dp[nt_te], act_te[nt_te])
            priv_g = clean_atk - m_ct['acc']
            u_loss = clean_util - nutil
            u_loss_aw = clean_util - nutil_aw

            rows.append(dict(
                epsilon=eps,
                ct_acc=m_ct['acc'],    ct_recall=m_ct['recall'],  ct_f1=m_ct['f1'],
                aw_acc=m_aw['acc'],    aw_recall=m_aw['recall'],  aw_f1=m_aw['f1'],
                bal_acc=m_bal['acc'],  bal_recall=m_bal['recall'], bal_f1=m_bal['f1'],
                priv_gain=priv_g, util_noisy=nutil, util_loss=u_loss,
                util_noisy_aw=nutil_aw, util_loss_aw=u_loss_aw,
                policy_mode='fixed',
            ))
            print(f"{eps:>6.2f}  "
                  f"{m_ct['acc']:>8.4f}  {m_ct['recall']:>8.4f}  {m_ct['f1']:>8.4f}  "
                  f"{m_aw['acc']:>8.4f}  {m_aw['recall']:>8.4f}  {m_aw['f1']:>8.4f}  "
                  f"{m_bal['acc']:>9.4f}  {m_bal['f1']:>8.4f}  "
                  f"{priv_g:>10.4f}  {nutil:>8.4f}  {nutil_aw:>8.4f}  {u_loss:>10.4f}")

    return rows


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# STAGE 2+3 â€” Keystroke100: DP Defense + Privacy Sweep
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def dp_eval_keystroke(res):
    print("\n" + "=" * 72)
    print("STAGE 2+3 â€” Keystroke100: Laplace DP  +  Privacy Sweep")
    print("=" * 72)

    X_tv, y_tv  = res['X_tv'], res['y_tv']
    X_te, y_te  = res['X_te'], res['y_te']
    clf, sc, scaled = res['best_clf'], res['scaler'], res['scaled']
    classes = res['classes']

    clean1  = res['baseline']['top1']
    clean5  = res['baseline']['top5']
    clean_f1  = res['baseline']['f1']
    clean_eer = res['baseline']['mean_eer']
    print(f"\nClean attack  Top-1={clean1:.4f}  Top-5={clean5:.4f}  "
          f"wt-F1={clean_f1:.4f}  EER={clean_eer:.4f}")
    print(f"Random-chance baseline: {1/len(classes):.4f}  (1/100 users)")

    hdr = (f"\n{'eps':>6}  {'ct-Top1':>9}  {'ct-F1':>8}  "
           f"{'aw-Top1':>9}  {'aw-Top5':>9}  {'aw-F1':>8}  {'aw-EER':>8}  "
           f"{'PrivGain':>10}")
    print(hdr)
    print("-" * 80)

    rows = []
    for eps in EPSILONS:
        X_te_dp = laplace_dp(X_tv, X_te, eps)
        te_in   = sc.transform(X_te_dp) if scaled else X_te_dp
        m_ct    = multiclass_metrics(clf, te_in, y_te, classes)

        X_tv_dp = laplace_dp(X_tv, X_tv, eps)
        sc_aw   = StandardScaler().fit(X_tv_dp)
        tv_aw   = sc_aw.transform(X_tv_dp) if scaled else X_tv_dp
        te_aw   = sc_aw.transform(X_te_dp) if scaled else X_te_dp
        clf_aw  = clf.__class__(**clf.get_params())
        clf_aw.fit(tv_aw, y_tv)
        m_aw    = multiclass_metrics(clf_aw, te_aw, y_te, classes)
        priv_g  = clean1 - m_ct['top1']

        rows.append(dict(
            epsilon=eps,
            ct_top1=m_ct['top1'], ct_f1=m_ct['f1'],
            aw_top1=m_aw['top1'], aw_top5=m_aw['top5'],
            aw_f1=m_aw['f1'],    aw_eer=m_aw['mean_eer'],
            priv_gain=priv_g,
        ))
        print(f"{eps:>6.2f}  "
              f"{m_ct['top1']:>9.4f}  {m_ct['f1']:>8.4f}  "
              f"{m_aw['top1']:>9.4f}  {m_aw['top5']:>9.4f}  "
              f"{m_aw['f1']:>8.4f}  {m_aw['mean_eer']:>8.4f}  "
              f"{priv_g:>10.4f}")

    return rows


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# MAIN
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# OUTPUT: Excel + Visuals
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

_ROOT    = os.path.dirname(os.path.abspath(__file__))
OUT_DIR  = os.path.join(_ROOT, 'output')
VIS_DIR  = os.path.join(_ROOT, 'visuals')
os.makedirs(OUT_DIR, exist_ok=True)
os.makedirs(VIS_DIR, exist_ok=True)


def _save_workbook_with_fallback(wb, path):
    """
    Save workbook to path, falling back to timestamped filename if locked.
    """
    try:
        wb.save(path)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        alt = f'{base}_{ts}{ext}'
        wb.save(alt)
        return alt

# â”€â”€ Style helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_HDR_FILL   = PatternFill('solid', fgColor='1F4E79')
_SUB_FILL   = PatternFill('solid', fgColor='2E75B6')
_ALT_FILL   = PatternFill('solid', fgColor='D6E4F0')
_WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')
_HDR_FONT   = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
_BODY_FONT  = Font(name='Calibri', size=10)
_BOLD_FONT  = Font(bold=True, name='Calibri', size=10)
_CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT       = Alignment(horizontal='left',   vertical='center')

def _thin_border():
    s = Side(border_style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)

def _style_header(cell, sub=False):
    cell.fill      = _SUB_FILL if sub else _HDR_FILL
    cell.font      = _HDR_FONT
    cell.alignment = _CENTER
    cell.border    = _thin_border()

def _style_body(cell, alt=False, bold=False):
    cell.fill      = _ALT_FILL if alt else _WHITE_FILL
    cell.font      = _BOLD_FONT if bold else _BODY_FONT
    cell.alignment = _CENTER
    cell.border    = _thin_border()

def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 28)

def _write_table(ws, headers, rows, start_row=1, alt_rows=True):
    """Write a styled table into worksheet ws starting at start_row."""
    for ci, h in enumerate(headers, 1):
        c = ws.cell(start_row, ci, h)
        _style_header(c)
    for ri, row in enumerate(rows, start_row + 1):
        alt = alt_rows and (ri % 2 == 0)
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            _style_body(c, alt=alt)
    return start_row + len(rows) + 1   # next free row


# â”€â”€ WISDM Excel â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def export_wisdm_excel(wisdm_r, wisdm_dp, wisdm_dp_fixed=None):
    wb   = openpyxl.Workbook()
    mode = wisdm_dp[0].get('policy_mode', 'fixed') if wisdm_dp else 'fixed'

    # â”€â”€ Sheet 1: Baseline (no DP) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws = wb.active
    ws.title = 'WISDM Baseline (No DP)'
    ws.sheet_view.showGridLines = False

    title = ws.cell(1, 1, 'WISDM â€” Typing Task Inference: Baseline Results (No Differential Privacy)')
    title.font      = Font(bold=True, name='Calibri', size=13, color='1F4E79')
    title.alignment = _LEFT
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 22

    note = ws.cell(2, 1,
        'Task: binary typing-related (F=typing, Q=writing) vs. all other smartwatch activities  |  '
        'Features: 182 (watch accel + gyro ARFF)  |  Split: subject-disjoint 70/15/15')
    note.font      = Font(italic=True, name='Calibri', size=9, color='595959')
    note.alignment = _LEFT
    ws.merge_cells('A2:G2')

    # --- All-models table (val + test) ---
    ws.cell(4, 1, 'All Models â€” Validation & Test').font = Font(bold=True, name='Calibri', size=11, color='1F4E79')

    hdrs = ['Model', 'Split', 'Accuracy', 'Precision', 'Recall', 'F1-Score', 'ROC-AUC']
    b    = wisdm_r['baseline']

    # Dynamic rows from stored per-model metrics
    model_rows = []
    for name, (m_vl, m_te, _) in wisdm_r['model_metrics'].items():
        model_rows.append((name, 'Validation',
            m_vl['acc'], m_vl['precision'], m_vl['recall'], m_vl['f1'], m_vl['roc_auc']))
        model_rows.append((name, 'Test',
            m_te['acc'], m_te['precision'], m_te['recall'], m_te['f1'], m_te['roc_auc']))
    formatted = [
        (m, sp) + tuple(f'{v:.4f}' for v in vals)
        for m, sp, *vals in model_rows
    ]
    next_row = _write_table(ws, hdrs, formatted, start_row=5)

    # --- Best model final test ---
    ws.cell(next_row + 1, 1, f'Best Model: {wisdm_r["best_name"]}  (selected on val F1; refitted on train+val)').font = \
        Font(bold=True, name='Calibri', size=11, color='1F4E79')
    best_row = [
        (wisdm_r['best_name'], 'Test (final)',
         f'{b["acc"]:.4f}', f'{b["precision"]:.4f}',
         f'{b["recall"]:.4f}', f'{b["f1"]:.4f}', f'{b["roc_auc"]:.4f}')
    ]
    _write_table(ws, hdrs, best_row, start_row=next_row + 2)
    _autofit(ws)

    # â”€â”€ Sheet 2: DP Sweep â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws2_name = 'WISDM ADP Sweep' if mode == 'adaptive' else 'WISDM Fixed-DP Sweep'
    ws2 = wb.create_sheet(ws2_name)
    ws2.sheet_view.showGridLines = False

    if mode == 'adaptive':
        title_txt = 'WISDM â€” Privacy-Utility Tradeoff: Adaptive Laplace DP (Typing-Triggered)'
    else:
        title_txt = 'WISDM â€” Privacy-Utility Tradeoff: Laplace DP Sweep (IBM diffprivlib)'
    t2 = ws2.cell(1, 1, title_txt)
    t2.font      = Font(bold=True, name='Calibri', size=13, color='1F4E79')
    t2.alignment = _LEFT
    ws2_cols = 15 if mode == 'adaptive' else 8
    ws2.merge_cells(f'A1:{get_column_letter(ws2_cols)}1')
    ws2.row_dimensions[1].height = 22

    if mode == 'adaptive':
        n2_txt = (
            f'Best attack model: {wisdm_r["best_name"]}  |  '
            f'Clean: acc={wisdm_r["baseline"]["acc"]:.4f}  '
            f'typing-recall={wisdm_r["baseline"]["recall"]:.4f}  '
            f'typing-F1={wisdm_r["baseline"]["f1"]:.4f}  |  '
            f'Mechanism: Adaptive Laplace ({ADP_POLICY}), tau={ADP_TAU:.2f}, '
            f'eps_non_typing={ADP_EPS_NONTYPING:.2f}, jitter_frac={ADP_JITTER_FRAC:.2f}')
    else:
        n2_txt = (
            f'Best attack model: {wisdm_r["best_name"]}  |  '
            f'Clean: acc={wisdm_r["baseline"]["acc"]:.4f}  '
            f'typing-recall={wisdm_r["baseline"]["recall"]:.4f}  '
            f'typing-F1={wisdm_r["baseline"]["f1"]:.4f}  |  '
            f'Mechanism: Laplace (IBM diffprivlib), per-feature  |  Bounds from training split')
    n2 = ws2.cell(2, 1, n2_txt)
    n2.font      = Font(italic=True, name='Calibri', size=9, color='595959')
    n2.alignment = _LEFT
    ws2.merge_cells(f'A2:{get_column_letter(ws2_cols)}2')

    if mode == 'adaptive':
        dp_hdrs = [
            'Typing Îµ',
            'Non-typing Îµ',
            'Mean Îµ on Test',
            'Std Îµ on Test',
            'Typing F1\n(clean-trained)',
            'Typing F1\n(attacker-aware)',
            'Typing F1\n(gate-only baseline)',
            'AWâ†”Gate Agreement',
            'AWâ†”Gate Corr',
            'Attack Acc\n(clean-trained)',
            'Privacy Gain\n(clean acc âˆ’ noisy acc)',
            'Utility Acc\n(clean-trained utility)',
            'Utility Acc\n(utility-aware)',
            'Typing Coverage\n(gate recall)',
            'Over-trigger Rate\n(gate FPR)',
        ]
        dp_rows = [
            (f'{r["eps_typing"]:.2f}', f'{r["eps_non_typing"]:.2f}',
             f'{r["eff_eps"]:.4f}', f'{r.get("eff_eps_std", float("nan")):.4f}',
             f'{r["ct_f1"]:.4f}', f'{r["aw_f1"]:.4f}', f'{r["gate_f1"]:.4f}',
             f'{r["aw_gate_agree"]:.4f}',
             f'{r["aw_gate_corr"]:.4f}' if np.isfinite(r["aw_gate_corr"]) else 'nan',
             f'{r["ct_acc"]:.4f}', f'{r["priv_gain"]:.4f}',
             f'{r["util_noisy"]:.4f}', f'{r["util_noisy_aw"]:.4f}',
             f'{r["gate_recall"]:.4f}', f'{r["gate_fpr"]:.4f}')
            for r in wisdm_dp
        ]
    else:
        dp_hdrs = [
            'Privacy Budget Îµ',
            'Attack Acc\n(clean-trained)',
            'Typing Recall\n(clean-trained)',
            'Typing F1\n(clean-trained)',
            'Typing F1\n(attacker-aware)',
            'Utility Acc\n(clean-trained utility)',
            'Utility Acc\n(utility-aware)',
            'Privacy Gain\n(clean acc âˆ’ noisy acc)',
        ]
        dp_rows = [
            (f'{r["epsilon"]:.2f}',
             f'{r["ct_acc"]:.4f}', f'{r["ct_recall"]:.4f}', f'{r["ct_f1"]:.4f}',
             f'{r["aw_f1"]:.4f}', f'{r["util_noisy"]:.4f}',
             f'{r["util_noisy_aw"]:.4f}', f'{r["priv_gain"]:.4f}')
            for r in wisdm_dp
        ]
    _write_table(ws2, dp_hdrs, dp_rows, start_row=4)
    _autofit(ws2)

    # Optional Sheet 3: fixed-DP baseline sweep (for ADP comparison in the same workbook)
    if mode == 'adaptive' and wisdm_dp_fixed:
        ws3 = wb.create_sheet('WISDM Fixed-DP Sweep')
        ws3.sheet_view.showGridLines = False

        t3 = ws3.cell(1, 1, 'WISDM â€” Privacy-Utility Tradeoff: Fixed Laplace DP Baseline')
        t3.font      = Font(bold=True, name='Calibri', size=13, color='1F4E79')
        t3.alignment = _LEFT
        ws3.merge_cells('A1:H1')
        ws3.row_dimensions[1].height = 22

        n3 = ws3.cell(
            2, 1,
            f'Best attack model: {wisdm_r["best_name"]}  |  '
            f'Clean: acc={wisdm_r["baseline"]["acc"]:.4f}  '
            f'typing-recall={wisdm_r["baseline"]["recall"]:.4f}  '
            f'typing-F1={wisdm_r["baseline"]["f1"]:.4f}  |  '
            f'Mechanism: Fixed Laplace (IBM diffprivlib), per-feature'
        )
        n3.font      = Font(italic=True, name='Calibri', size=9, color='595959')
        n3.alignment = _LEFT
        ws3.merge_cells('A2:H2')

        fixed_hdrs = [
            'Privacy Budget Îµ',
            'Attack Acc\n(clean-trained)',
            'Typing Recall\n(clean-trained)',
            'Typing F1\n(clean-trained)',
            'Typing F1\n(attacker-aware)',
            'Utility Acc\n(clean-trained utility)',
            'Utility Acc\n(utility-aware)',
            'Privacy Gain\n(clean acc âˆ’ noisy acc)',
        ]
        fixed_rows = [
            (f'{r["epsilon"]:.2f}',
             f'{r["ct_acc"]:.4f}', f'{r["ct_recall"]:.4f}', f'{r["ct_f1"]:.4f}',
             f'{r["aw_f1"]:.4f}', f'{r["util_noisy"]:.4f}',
             f'{r["util_noisy_aw"]:.4f}', f'{r["priv_gain"]:.4f}')
            for r in wisdm_dp_fixed
        ]
        _write_table(ws3, fixed_hdrs, fixed_rows, start_row=4)
        _autofit(ws3)

    fname = 'wisdm_results_adp.xlsx' if mode == 'adaptive' else 'wisdm_results_fixed_dp.xlsx'
    path = os.path.join(OUT_DIR, fname)
    saved = _save_workbook_with_fallback(wb, path)
    print(f"  Saved: {saved}")


# â”€â”€ Keystroke100 Excel â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def export_keystroke_excel(ks_r, ks_dp):
    wb = openpyxl.Workbook()

    # â”€â”€ Sheet 1: Baseline â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws = wb.active
    ws.title = 'Keystroke100 Baseline (No DP)'
    ws.sheet_view.showGridLines = False

    t = ws.cell(1, 1, 'Keystroke100 â€” Closed-Set User Identification: Baseline Results (No Differential Privacy)')
    t.font      = Font(bold=True, name='Calibri', size=13, color='1F4E79')
    t.alignment = _LEFT
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 22

    n = ws.cell(2, 1,
        'Task: 100-class user identification  |  Features: 13 (7 latency + 6 pressure-summary)  |  '
        'Split: stratified 70/15/15  |  Password: "try4-mbs" (all users)')
    n.font      = Font(italic=True, name='Calibri', size=9, color='595959')
    n.alignment = _LEFT
    ws.merge_cells('A2:H2')

    ws.cell(4, 1, 'All Models â€” Validation & Test').font = \
        Font(bold=True, name='Calibri', size=11, color='1F4E79')

    hdrs = ['Model', 'Split', 'Top-1 Acc', 'Top-5 Acc', 'Precision', 'Recall', 'F1-Score', 'Mean EER']
    model_rows = []
    for name, (m_vl, m_te, _) in ks_r['model_metrics'].items():
        model_rows.append((name, 'Validation',
            m_vl['top1'], m_vl['top5'], m_vl['precision'], m_vl['recall'], m_vl['f1'], m_vl['mean_eer']))
        model_rows.append((name, 'Test',
            m_te['top1'], m_te['top5'], m_te['precision'], m_te['recall'], m_te['f1'], m_te['mean_eer']))
    formatted = [
        (m, sp) + tuple(f'{v:.4f}' for v in vals)
        for m, sp, *vals in model_rows
    ]
    next_row = _write_table(ws, hdrs, formatted, start_row=5)

    b = ks_r['baseline']
    ws.cell(next_row + 1, 1,
            f'Best Model: {ks_r["best_name"]}  (selected on val Top-1; refitted on train+val)').font = \
        Font(bold=True, name='Calibri', size=11, color='1F4E79')
    best_row = [(
        ks_r['best_name'], 'Test (final)',
        f'{b["top1"]:.4f}', f'{b["top5"]:.4f}',
        f'{b["precision"]:.4f}', f'{b["recall"]:.4f}',
        f'{b["f1"]:.4f}', f'{b["mean_eer"]:.4f}',
    )]
    _write_table(ws, hdrs, best_row, start_row=next_row + 2)
    _autofit(ws)

    # â”€â”€ Sheet 2: DP Sweep â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws2 = wb.create_sheet('Keystroke100 Fixed-DP Sweep')
    ws2.sheet_view.showGridLines = False

    t2 = ws2.cell(1, 1, 'Keystroke100 â€” Privacy Sweep: Fixed Laplace DP (IBM diffprivlib)')
    t2.font      = Font(bold=True, name='Calibri', size=13, color='1F4E79')
    t2.alignment = _LEFT
    ws2.merge_cells('A1:E1')
    ws2.row_dimensions[1].height = 22

    n2 = ws2.cell(2, 1,
        f'Best attack model: {ks_r["best_name"]}  |  '
        f'Clean Top-1: {ks_r["baseline"]["top1"]:.4f}  |  '
        f'Clean Top-5: {ks_r["baseline"]["top5"]:.4f}  |  '
        f'Mechanism: Laplace, per-feature  |  Bounds from training split')
    n2.font      = Font(italic=True, name='Calibri', size=9, color='595959')
    n2.alignment = _LEFT
    ws2.merge_cells('A2:E2')

    dp_hdrs = [
        'Privacy Budget Îµ',
        'Attack Top-1\n(clean-trained)',
        'Attack Top-1\n(attacker-aware)',
        'Attack EER\n(attacker-aware)',
        'Privacy Gain\n(clean Top-1 âˆ’ noisy Top-1)',
    ]
    dp_rows = [
        (f'{r["epsilon"]:.2f}',
         f'{r["ct_top1"]:.4f}',
         f'{r["aw_top1"]:.4f}',
         f'{r["aw_eer"]:.4f}',
         f'{r["priv_gain"]:.4f}')
        for r in ks_dp
    ]
    _write_table(ws2, dp_hdrs, dp_rows, start_row=4)
    _autofit(ws2)

    path = os.path.join(OUT_DIR, 'keystroke100_results_fixed_dp.xlsx')
    saved = _save_workbook_with_fallback(wb, path)
    print(f"  Saved: {saved}")


# â”€â”€ Visuals â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
_PALETTE = ['#1F4E79', '#2E75B6', '#70AD47', '#ED7D31', '#FF0000']
_STYLE = dict(dpi=150, bbox_inches='tight')

def _save(fig, name):
    p = os.path.join(VIS_DIR, name)
    fig.savefig(p, **_STYLE)
    plt.close(fig)
    print(f"  Saved: {p}")


def plot_wisdm_baseline():
    """Bar chart: WISDM model comparison across metrics (test set)."""
    models  = ['Logistic\nRegression', 'Random\nForest', 'SVM\n(RBF)']
    metrics = ['Accuracy', 'Precision', 'Recall', 'F1-Score', 'ROC-AUC']
    data = [
        [0.9606, 0.8740, 0.7500, 0.8073, 0.9356],
        [0.9576, 0.9691, 0.6351, 0.7673, 0.9648],
        [0.9624, 0.9185, 0.7230, 0.8091, 0.9444],
    ]

    x   = np.arange(len(metrics))
    w   = 0.22
    fig, ax = plt.subplots(figsize=(11, 5))
    for i, (m, vals) in enumerate(zip(models, data)):
        bars = ax.bar(x + i * w, vals, w, label=m,
                      color=_PALETTE[i], edgecolor='white', linewidth=0.6)
        for b in bars:
            ax.text(b.get_x() + b.get_width() / 2, b.get_height() + 0.005,
                    f'{b.get_height():.3f}', ha='center', va='bottom',
                    fontsize=7.5, color='#333333')

    ax.set_title('WISDM â€” Typing Task Inference: Baseline Model Comparison (Test Set)',
                 fontsize=12, fontweight='bold', pad=12)
    ax.set_ylabel('Score', fontsize=10)
    ax.set_xticks(x + w)
    ax.set_xticklabels(metrics, fontsize=10)
    ax.set_ylim(0.55, 1.08)
    ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.2f'))
    ax.legend(title='Model', fontsize=9, title_fontsize=9)
    ax.spines[['top', 'right']].set_visible(False)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    _save(fig, 'wisdm_baseline_comparison.png')


def plot_keystroke_baseline():
    """Bar chart: Keystroke100 model comparison (test set)."""
    models  = ['kNN', 'Random\nForest', 'SVM\n(RBF)']
    metrics = ['Top-1 Acc', 'Top-5 Acc', 'Precision', 'Recall', 'F1-Score']
    data = [
        [0.6267, 0.8400, 0.6159, 0.6267, 0.5941],
        [0.9600, 1.0000, 0.9689, 0.9600, 0.9573],
        [0.6800, 0.8667, 0.6456, 0.6800, 0.6420],
    ]

    x = np.arange(len(metrics))
    w = 0.22
    fig, ax = plt.subplots(figsize=(11, 5))
    for i, (m, vals) in enumerate(zip(models, data)):
        bars = ax.bar(x + i * w, vals, w, label=m,
                      color=_PALETTE[i], edgecolor='white', linewidth=0.6)
        for b in bars:
            ax.text(b.get_x() + b.get_width() / 2, b.get_height() + 0.005,
                    f'{b.get_height():.3f}', ha='center', va='bottom',
                    fontsize=7.5, color='#333333')

    ax.set_title('Keystroke100 â€” User Identification: Baseline Model Comparison (Test Set)',
                 fontsize=12, fontweight='bold', pad=12)
    ax.set_ylabel('Score', fontsize=10)
    ax.set_xticks(x + w)
    ax.set_xticklabels(metrics, fontsize=10)
    ax.set_ylim(0.45, 1.12)
    ax.yaxis.set_major_formatter(mticker.FormatStrFormatter('%.2f'))
    ax.legend(title='Model', fontsize=9, title_fontsize=9)
    ax.spines[['top', 'right']].set_visible(False)
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    fig.tight_layout()
    _save(fig, 'keystroke_baseline_comparison.png')


def plot_wisdm_dp_tradeoff(wisdm_dp):
    """WISDM: privacy-utility tradeoff line plot."""
    mode = wisdm_dp[0].get('policy_mode', 'fixed') if wisdm_dp else 'fixed'
    eps       = [r['epsilon']    for r in wisdm_dp]
    atk_ct    = [r['ct_acc']     for r in wisdm_dp]
    atk_aw    = [r['aw_acc']     for r in wisdm_dp]
    priv_gain = [r['priv_gain']  for r in wisdm_dp]
    util      = [r['util_noisy'] for r in wisdm_dp]
    util_loss = [r['util_loss']  for r in wisdm_dp]

    fig, axes = plt.subplots(1, 2, figsize=(13, 5))
    if mode == 'adaptive':
        fig.suptitle('WISDM â€” Privacy-Utility Tradeoff under Adaptive Laplace DP',
                     fontsize=13, fontweight='bold')
        xlab = 'Typing Privacy Budget Îµ_t  (log scale; Îµ_n fixed)'
    else:
        fig.suptitle('WISDM â€” Privacy-Utility Tradeoff under Laplace DP',
                     fontsize=13, fontweight='bold')
        xlab = 'Privacy Budget Îµ  (log scale)'

    # Left: attack accuracy vs epsilon
    ax = axes[0]
    ax.plot(eps, atk_ct,  'o-',  color=_PALETTE[0], lw=2, ms=7,
            label='Attack (clean-trained â†’ noisy)')
    ax.plot(eps, atk_aw,  's--', color=_PALETTE[4], lw=2, ms=7,
            label='Attack (attacker-aware, standard)')
    ax.axhline(0.9602, color='#333', lw=1.2, ls=':', label='Clean baseline (0.9602)')
    ax.set_xscale('log')
    ax.set_xlabel(xlab, fontsize=10)
    ax.set_ylabel('Attack Accuracy', fontsize=10)
    ax.set_title('Attack Accuracy vs. Îµ', fontsize=11)
    ax.legend(fontsize=7.5)
    ax.grid(True, linestyle='--', alpha=0.4)
    ax.spines[['top', 'right']].set_visible(False)
    ax.set_ylim(0, 1.05)

    # Right: utility and privacy gain vs epsilon
    ax2 = axes[1]
    ax2.plot(eps, util,      'o-',  color=_PALETTE[1], lw=2, ms=7, label='Utility Acc (noisy non-typing)')
    ax2.plot(eps, priv_gain, 's--', color=_PALETTE[3], lw=2, ms=7, label='Privacy Gain')
    ax2.plot(eps, util_loss, '^:',  color=_PALETTE[4], lw=2, ms=7, label='Utility Loss')
    ax2.set_xscale('log')
    ax2.set_xlabel(xlab, fontsize=10)
    ax2.set_ylabel('Score', fontsize=10)
    ax2.set_title('Utility & Privacy Gain vs. Îµ', fontsize=11)
    ax2.legend(fontsize=8.5)
    ax2.grid(True, linestyle='--', alpha=0.4)
    ax2.spines[['top', 'right']].set_visible(False)
    ax2.set_ylim(-0.05, 1.05)

    fig.tight_layout()
    _save(fig, 'wisdm_dp_tradeoff.png')


def plot_keystroke_dp(ks_dp):
    """Keystroke100: attack accuracy vs epsilon."""
    eps    = [r['epsilon']  for r in ks_dp]
    ct1    = [r['ct_top1']  for r in ks_dp]
    aw1    = [r['aw_top1']  for r in ks_dp]
    aw5    = [r['aw_top5']  for r in ks_dp]
    priv_g = [r['priv_gain'] for r in ks_dp]

    fig, axes = plt.subplots(1, 2, figsize=(13, 5))
    fig.suptitle('Keystroke100 â€” Privacy Degradation under Laplace DP',
                 fontsize=13, fontweight='bold')

    ax = axes[0]
    ax.plot(eps, ct1, 'o-', color=_PALETTE[0], lw=2, ms=7,
            label='Atk Top-1 (clean-trained â†’ noisy)')
    ax.plot(eps, aw1, 's--', color=_PALETTE[2], lw=2, ms=7,
            label='Atk Top-1 (attacker-aware)')
    ax.plot(eps, aw5, '^:', color=_PALETTE[1], lw=2, ms=7,
            label='Atk Top-5 (attacker-aware)')
    ax.axhline(0.9467, color='#888', lw=1.2, ls=':', label='Clean Top-1 baseline (0.9467)')
    ax.set_xscale('log')
    ax.set_xlabel('Privacy Budget Îµ  (log scale)', fontsize=10)
    ax.set_ylabel('Accuracy', fontsize=10)
    ax.set_title('Attack Accuracy vs. Îµ', fontsize=11)
    ax.legend(fontsize=8.5)
    ax.grid(True, linestyle='--', alpha=0.4)
    ax.spines[['top', 'right']].set_visible(False)
    ax.set_ylim(-0.05, 1.05)

    ax2 = axes[1]
    ax2.bar([str(e) for e in eps], priv_g,
            color=_PALETTE[0], edgecolor='white', linewidth=0.6, width=0.5)
    for xi, (e, pg) in enumerate(zip(eps, priv_g)):
        ax2.text(xi, pg + 0.005, f'{pg:.4f}', ha='center', va='bottom',
                 fontsize=9, color='#1F4E79', fontweight='bold')
    ax2.set_xlabel('Privacy Budget Îµ', fontsize=10)
    ax2.set_ylabel('Privacy Gain (clean Top-1 âˆ’ noisy Top-1)', fontsize=10)
    ax2.set_title('Privacy Gain vs. Îµ', fontsize=11)
    ax2.set_ylim(0, 1.08)
    ax2.grid(axis='y', linestyle='--', alpha=0.4)
    ax2.spines[['top', 'right']].set_visible(False)

    fig.tight_layout()
    _save(fig, 'keystroke_dp_privacy.png')


def plot_wisdm_adp_gate_diagnostics(wisdm_dp):
    """
    Adaptive-only diagnostics:
    - whether attacker-aware tracks gate artifacts,
    - how gate quality and utility evolve vs eps_t.
    """
    if not wisdm_dp or wisdm_dp[0].get('policy_mode') != 'adaptive':
        return

    eps = [r['epsilon'] for r in wisdm_dp]
    ct_f1 = [r['ct_f1'] for r in wisdm_dp]
    aw_f1 = [r['aw_f1'] for r in wisdm_dp]
    gate_f1 = [r.get('gate_f1', float('nan')) for r in wisdm_dp]
    aw_gate_agree = [r.get('aw_gate_agree', float('nan')) for r in wisdm_dp]
    aw_gate_corr = [r.get('aw_gate_corr', float('nan')) for r in wisdm_dp]
    gate_rec = [r.get('gate_recall', float('nan')) for r in wisdm_dp]
    gate_fpr = [r.get('gate_fpr', float('nan')) for r in wisdm_dp]
    util = [r['util_noisy'] for r in wisdm_dp]

    fig, axes = plt.subplots(1, 2, figsize=(13, 5))
    fig.suptitle('WISDM Adaptive DP â€” Gate Artifact Diagnostics',
                 fontsize=13, fontweight='bold')

    ax = axes[0]
    ax.plot(eps, ct_f1, 'o-', color=_PALETTE[0], lw=2, ms=7, label='Clean-trained typing F1')
    ax.plot(eps, aw_f1, 's--', color=_PALETTE[4], lw=2, ms=7, label='Attacker-aware typing F1')
    ax.plot(eps, gate_f1, 'd-.', color=_PALETTE[2], lw=2, ms=7, label='Gate-only typing F1')
    ax.plot(eps, aw_gate_agree, 'x:', color=_PALETTE[3], lw=2, ms=8, label='AW↔Gate agreement')
    ax.plot(eps, aw_gate_corr, 'p:', color='#6A5ACD', lw=2, ms=7, label='AW↔Gate corr')
    ax.set_xscale('log')
    ax.set_xlabel('Typing Privacy Budget Îµ_t  (log scale)', fontsize=10)
    ax.set_ylabel('Score', fontsize=10)
    ax.set_title('Attack vs. Gate Alignment', fontsize=11)
    ax.set_ylim(-0.02, 1.05)
    ax.legend(fontsize=8)
    ax.grid(True, linestyle='--', alpha=0.4)
    ax.spines[['top', 'right']].set_visible(False)

    ax2 = axes[1]
    ax2.plot(eps, util, 'o-', color=_PALETTE[1], lw=2, ms=7, label='Utility acc (non-typing)')
    ax2.plot(eps, gate_rec, 's--', color=_PALETTE[0], lw=2, ms=7, label='Gate recall')
    ax2.plot(eps, gate_fpr, 'd-.', color=_PALETTE[3], lw=2, ms=7, label='Gate FPR')
    ax2.set_xscale('log')
    ax2.set_xlabel('Typing Privacy Budget Îµ_t  (log scale)', fontsize=10)
    ax2.set_ylabel('Score', fontsize=10)
    ax2.set_title('Utility and Gate Quality', fontsize=11)
    ax2.set_ylim(-0.02, 1.05)
    ax2.legend(fontsize=8)
    ax2.grid(True, linestyle='--', alpha=0.4)
    ax2.spines[['top', 'right']].set_visible(False)

    fig.tight_layout()
    _save(fig, 'wisdm_adp_gate_diagnostics.png')


def plot_wisdm_adp_effective_eps(wisdm_dp):
    """Adaptive-only: effective epsilon distribution summary across eps_t sweep."""
    if not wisdm_dp or wisdm_dp[0].get('policy_mode') != 'adaptive':
        return

    eps_t = [r['epsilon'] for r in wisdm_dp]
    eff_mu = [r.get('eff_eps', float('nan')) for r in wisdm_dp]
    eff_sd = [r.get('eff_eps_std', float('nan')) for r in wisdm_dp]
    gate_rec = [r.get('gate_recall', float('nan')) for r in wisdm_dp]
    gate_fpr = [r.get('gate_fpr', float('nan')) for r in wisdm_dp]

    fig, ax = plt.subplots(figsize=(8.8, 5))
    ax.errorbar(eps_t, eff_mu, yerr=eff_sd, fmt='o-', color=_PALETTE[0], lw=2, ms=7,
                capsize=4, label='Effective ε (mean ± std)')
    ax.set_xscale('log')
    ax.set_xlabel('Typing Privacy Budget ε_t  (log scale)', fontsize=10)
    ax.set_ylabel('Effective ε on Test', fontsize=10)
    ax.set_title('WISDM Adaptive DP — Effective ε Distribution',
                 fontsize=12, fontweight='bold')
    ax.grid(True, linestyle='--', alpha=0.4)
    ax.spines[['top', 'right']].set_visible(False)

    ax2 = ax.twinx()
    ax2.plot(eps_t, gate_rec, 's--', color=_PALETTE[1], lw=1.8, ms=6, label='Gate recall')
    ax2.plot(eps_t, gate_fpr, 'd:', color=_PALETTE[3], lw=1.8, ms=6, label='Gate FPR')
    ax2.set_ylabel('Gate Metrics', fontsize=10)
    ax2.set_ylim(-0.02, 1.02)

    h1, l1 = ax.get_legend_handles_labels()
    h2, l2 = ax2.get_legend_handles_labels()
    ax.legend(h1 + h2, l1 + l2, fontsize=8.5, loc='best')

    fig.tight_layout()
    _save(fig, 'wisdm_adp_effective_eps.png')


def plot_combined_summary(wisdm_dp, ks_dp):
    """Privacy gain trends for both datasets on their own epsilon axes."""
    if not wisdm_dp or not ks_dp:
        return
    mode = wisdm_dp[0].get('policy_mode', 'fixed')
    eps_w = [r['epsilon'] for r in wisdm_dp]
    eps_k = [r['epsilon'] for r in ks_dp]
    wg    = [r['priv_gain'] for r in wisdm_dp]
    kg    = [r['priv_gain'] for r in ks_dp]

    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(eps_w, wg, 'o-', color=_PALETTE[0], lw=2, ms=7, label='WISDM (Typing Task)')
    ax.plot(eps_k, kg, 's--', color=_PALETTE[2], lw=2, ms=7, label='Keystroke100 (User ID)')
    for xw, yw in zip(eps_w, wg):
        ax.text(xw, yw + 0.01, f'{yw:.3f}', ha='center', va='bottom', fontsize=8)
    for xk, yk in zip(eps_k, kg):
        ax.text(xk, yk + 0.01, f'{yk:.3f}', ha='center', va='bottom', fontsize=8)

    ax.set_xscale('log')
    if mode == 'adaptive':
        xlab = 'Privacy Budget (WISDM uses typing Îµ_t; Keystroke100 uses fixed Îµ)'
    else:
        xlab = 'Privacy Budget Îµ'
    ax.set_xlabel(xlab, fontsize=10)
    ax.set_ylabel('Privacy Gain', fontsize=10)
    ax.set_title('Privacy Gain Trends â€” WISDM vs. Keystroke100',
                 fontsize=12, fontweight='bold')
    ax.set_ylim(0, 1.12)
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.4)
    ax.spines[['top', 'right']].set_visible(False)
    fig.tight_layout()
    _save(fig, 'combined_privacy_gain.png')


def plot_wisdm_balanced_variant(wisdm_dp):
    """
    Supplementary: side-by-side comparison of standard vs balanced attacker-aware across epsilon.
    """
    eps     = [r['epsilon']  for r in wisdm_dp]
    aw_acc  = [r['aw_acc']   for r in wisdm_dp]
    bal_acc = [r['bal_acc']  for r in wisdm_dp]
    aw_f1   = [r['aw_f1']    for r in wisdm_dp]
    bal_f1  = [r['bal_f1']   for r in wisdm_dp]
    ct_acc  = [r['ct_acc']   for r in wisdm_dp]
    ct_f1   = [r['ct_f1']    for r in wisdm_dp]

    fig, axes = plt.subplots(1, 2, figsize=(13, 5))
    fig.suptitle('WISDM â€” Standard vs. Balanced Attacker-Aware (Supplementary)',
                 fontsize=12, fontweight='bold')

    # Left: accuracy
    ax = axes[0]
    ax.plot(eps, ct_acc,  'o-',  color=_PALETTE[0], lw=2, ms=7,
            label='Clean-trained')
    ax.plot(eps, aw_acc,  's--', color=_PALETTE[4], lw=2, ms=7,
            label='Attacker-aware (standard)')
    ax.plot(eps, bal_acc, '^:',  color=_PALETTE[2], lw=2, ms=7,
            label='Attacker-aware (balanced)')
    ax.set_xscale('log')
    ax.set_xlabel('Privacy Budget Îµ  (log scale)', fontsize=10)
    ax.set_ylabel('Attack Accuracy', fontsize=10)
    ax.set_title('Accuracy: Standard vs. Balanced AW', fontsize=11)
    ax.legend(fontsize=8)
    ax.grid(True, linestyle='--', alpha=0.4)
    ax.spines[['top', 'right']].set_visible(False)
    ax.set_ylim(0.5, 1.02)

    # Right: typing-class F1
    ax2 = axes[1]
    ax2.plot(eps, ct_f1,  'o-',  color=_PALETTE[0], lw=2, ms=7,
             label='Clean-trained F1')
    ax2.plot(eps, aw_f1,  's--', color=_PALETTE[4], lw=2, ms=7,
             label='Attacker-aware F1 (standard)')
    ax2.plot(eps, bal_f1, '^:',  color=_PALETTE[2], lw=2, ms=7,
             label='Attacker-aware F1 (balanced)')
    ax2.axhline(0.0, color='#aaa', lw=1, ls=':')
    ax2.set_xscale('log')
    ax2.set_xlabel('Privacy Budget Îµ  (log scale)', fontsize=10)
    ax2.set_ylabel('Typing-Class F1-Score', fontsize=10)
    ax2.set_title('Typing F1: Standard vs. Balanced AW', fontsize=11)
    ax2.legend(fontsize=8)
    ax2.grid(True, linestyle='--', alpha=0.4)
    ax2.spines[['top', 'right']].set_visible(False)
    ax2.set_ylim(-0.05, 1.0)

    fig.tight_layout()
    _save(fig, 'wisdm_balanced_variant.png')


def plot_wisdm_attack_vs_utility(wisdm_dp):
    """Scatter: attack accuracy vs utility â€” shows the core tradeoff."""
    mode = wisdm_dp[0].get('policy_mode', 'fixed') if wisdm_dp else 'fixed'
    atk = [r['ct_acc']     for r in wisdm_dp]
    uti = [r['util_noisy'] for r in wisdm_dp]
    eps = [r['epsilon']    for r in wisdm_dp]

    fig, ax = plt.subplots(figsize=(7, 5))
    sc = ax.scatter(atk, uti, c=np.log10(eps), cmap='RdYlGn_r',
                    s=120, zorder=3, edgecolors='#333', linewidths=0.6)
    cbar = fig.colorbar(sc, ax=ax)
    cbar.set_label('logâ‚â‚€(Îµ)', fontsize=9)
    cbar.set_ticks(np.log10(eps))
    cbar.set_ticklabels([str(e) for e in eps])

    for x_pt, y_pt, e in zip(atk, uti, eps):
        lbl = f'Îµ_t={e}' if mode == 'adaptive' else f'Îµ={e}'
        ax.annotate(lbl, (x_pt, y_pt),
                    textcoords='offset points', xytext=(6, 4), fontsize=8.5)

    # Clean baseline point
    ax.scatter([0.9602], [0.6518], marker='*', s=200, color='gold',
               edgecolors='#333', linewidths=0.8, zorder=4, label='Clean baseline')
    ax.annotate('Clean', (0.9602, 0.6518), textcoords='offset points',
                xytext=(6, -12), fontsize=8.5, color='#8B6914')

    ax.set_xlabel('Attack Accuracy (clean-trained â†’ DP-noisy test)', fontsize=10)
    ax.set_ylabel('Utility Accuracy (non-typing, DP-noisy test)', fontsize=10)
    if mode == 'adaptive':
        ttl = ('WISDM â€” Attack vs. Utility Tradeoff (Adaptive DP)\n'
               '(ideal: low attack, high utility â†’ top-left)')
    else:
        ttl = ('WISDM â€” Attack vs. Utility Tradeoff\n'
               '(ideal: low attack, high utility â†’ top-left)')
    ax.set_title(ttl, fontsize=11, fontweight='bold')
    ax.legend(fontsize=9)
    ax.grid(True, linestyle='--', alpha=0.4)
    ax.spines[['top', 'right']].set_visible(False)
    fig.tight_layout()
    _save(fig, 'wisdm_attack_vs_utility.png')


def generate_all_outputs(wisdm_r, ks_r, wisdm_dp, ks_dp, wisdm_dp_fixed=None):
    print("\n" + "=" * 72)
    print("Exporting Excel tables â†’ output/")
    print("=" * 72)
    export_wisdm_excel(wisdm_r, wisdm_dp, wisdm_dp_fixed=wisdm_dp_fixed)
    export_keystroke_excel(ks_r, ks_dp)

    print("\n" + "=" * 72)
    print("Generating visuals â†’ visuals/")
    print("=" * 72)
    plot_wisdm_baseline()
    plot_keystroke_baseline()
    plot_wisdm_dp_tradeoff(wisdm_dp)
    plot_wisdm_adp_gate_diagnostics(wisdm_dp)
    plot_wisdm_adp_effective_eps(wisdm_dp)
    plot_keystroke_dp(ks_dp)
    plot_combined_summary(wisdm_dp, ks_dp)
    plot_wisdm_attack_vs_utility(wisdm_dp)
    plot_wisdm_balanced_variant(wisdm_dp)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# MAIN
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

if __name__ == '__main__':
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    validate_input_paths()

    print("=" * 72)
    print("Behavioral Biometric Analysis Pipeline")
    print("WISDM: typing-task inference  |  Keystroke100: user identification")
    print(f"DP strategy (WISDM): {DP_STRATEGY}")
    print("=" * 72)

    wisdm_r  = stage1_wisdm()
    ks_r     = stage1_keystroke()

    wisdm_dp = dp_eval_wisdm(wisdm_r, mode=DP_STRATEGY)
    wisdm_dp_fixed = dp_eval_wisdm(wisdm_r, mode='fixed') if DP_STRATEGY == 'adaptive' else None
    ks_dp    = dp_eval_keystroke(ks_r)

    generate_all_outputs(wisdm_r, ks_r, wisdm_dp, ks_dp, wisdm_dp_fixed=wisdm_dp_fixed)

    print("\n" + "=" * 72)
    print("PIPELINE COMPLETE")
    print("=" * 72)
